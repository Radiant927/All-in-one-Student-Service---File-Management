"""统计接口：首页仪表盘数据、趋势图、Excel 导出"""
from datetime import datetime, date, timedelta
from typing import Optional, List
from io import BytesIO
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models import User, Transfer, TransferStatus, FileType, Urgency, Campus
from app.schemas import DashboardStats
from app.auth import get_current_user

router = APIRouter(prefix="/api/stats", tags=["统计与导出"])

@router.get("/dashboard", response_model=DashboardStats, summary="首页仪表盘统计")
def dashboard_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    pending_receive = (
        db.query(Transfer)
        .filter(
            Transfer.to_campus == current_user.campus,
            Transfer.status.in_([TransferStatus.PENDING, TransferStatus.OVERDUE]),
        )
        .count()
    )
    my_pending_confirm = (
        db.query(Transfer)
        .filter(
            Transfer.created_by == current_user.id,
            Transfer.status.in_([TransferStatus.PENDING, TransferStatus.OVERDUE]),
        )
        .count()
    )
    now = datetime.now()
    month_start = datetime(now.year, now.month, 1)
    total_this_month = db.query(Transfer).filter(Transfer.created_at >= month_start).count()
    confirmed_this_month = (
        db.query(Transfer)
        .filter(Transfer.created_at >= month_start, Transfer.status == TransferStatus.CONFIRMED)
        .count()
    )
    return {
        "pending_receive": pending_receive,
        "my_pending_confirm": my_pending_confirm,
        "total_this_month": total_this_month,
        "confirmed_this_month": confirmed_this_month,
    }

@router.get("/trend", summary="近 7 天转交趋势")
def transfer_trend(
    days: int = Query(7, ge=1, le=30, description="统计天数"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    today = date.today()
    date_list = [today - timedelta(days=i) for i in range(days - 1, -1, -1)]
    result = []
    for d in date_list:
        day_start = datetime.combine(d, datetime.min.time())
        day_end = datetime.combine(d, datetime.max.time())
        day_total = db.query(Transfer).filter(Transfer.created_at >= day_start, Transfer.created_at <= day_end).count()
        day_confirmed = (
            db.query(Transfer)
            .filter(Transfer.confirm_time >= day_start, Transfer.confirm_time <= day_end, Transfer.status == TransferStatus.CONFIRMED)
            .count()
        )
        result.append({"date": d.strftime("%Y-%m-%d"), "total": day_total, "confirmed": day_confirmed})
    return result

@router.get("/export", summary="导出转交单 Excel")
def export_transfers(
    status: Optional[TransferStatus] = None,
    file_type: Optional[FileType] = None,
    urgency: Optional[Urgency] = None,
    role: Optional[str] = Query(None, description="sent=我发起的, received=发给我的"),
    keyword: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.services.transfer_service import get_transfer_list
    items, total = get_transfer_list(
        db=db, user=current_user, page=1, page_size=10000,
        status=status, file_type=file_type, urgency=urgency,
        role=role, keyword=keyword, date_from=date_from, date_to=date_to,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "转交单列表"
    headers = [
        "编号", "标题", "状态", "紧急程度", "文件类型",
        "出发校区", "目的校区", "转交同学", "接收人",
        "发车时间", "预计到达", "确认时间", "创建时间",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    status_map = {"pending": "待接收", "confirmed": "已确认", "overdue": "已逾期", "exception": "异常", "cancelled": "已撤回"}
    urgency_map = {"normal": "普通", "urgent": "加急", "critical": "特急"}
    filetype_map = {"admin": "行政文件", "teaching": "教学资料", "student": "学生材料", "finance": "财务票据", "other": "其他"}
    campus_map = {"nanhai": "南海校区", "shipai": "石牌校区"}
    for row, t in enumerate(items, 2):
        ws.cell(row=row, column=1, value=t.transfer_no)
        ws.cell(row=row, column=2, value=t.title)
        ws.cell(row=row, column=3, value=status_map.get(t.status.value, t.status.value))
        ws.cell(row=row, column=4, value=urgency_map.get(t.urgency.value, t.urgency.value))
        ws.cell(row=row, column=5, value=filetype_map.get(t.file_type.value, t.file_type.value))
        ws.cell(row=row, column=6, value=campus_map.get(t.from_campus.value, t.from_campus.value))
        ws.cell(row=row, column=7, value=campus_map.get(t.to_campus.value, t.to_campus.value))
        ws.cell(row=row, column=8, value=t.courier_name)
        ws.cell(row=row, column=9, value=t.receiver_name)
        ws.cell(row=row, column=10, value=t.depart_time.strftime("%Y-%m-%d %H:%M") if t.depart_time else "")
        ws.cell(row=row, column=11, value=t.estimate_arrive_time.strftime("%Y-%m-%d %H:%M") if t.estimate_arrive_time else "")
        ws.cell(row=row, column=12, value=t.confirm_time.strftime("%Y-%m-%d %H:%M") if t.confirm_time else "")
        ws.cell(row=row, column=13, value=t.created_at.strftime("%Y-%m-%d %H:%M:%S"))
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from urllib.parse import quote
    filename = f"转交单列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )